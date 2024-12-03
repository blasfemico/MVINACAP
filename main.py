import asyncio
import os
import random
import time
import re
from datetime import datetime
from typing import List, Generator
import cv2
import easyocr
import requests
from flask import Flask, render_template, Response, request, redirect, url_for, jsonify, flash
from openpyxl import Workbook, load_workbook
import pytesseract
import numpy as np 
from pynput import keyboard
from typing import Tuple
import logging
import openpyxl
from CCTV.sistema_gestion import SistemaGestion
from CCTV.trainer import getImagesAndLabels
from qreader import QReader


def photos_count():
    path = 'CCTV/fotos'
    imagePaths = [os.path.join(path,f) for f in os.listdir(path)]
    return len(imagePaths)

def train(recognizer):
    print ("Training faces. It will take a few seconds. Wait ......")
    faces,ids = getImagesAndLabels('CCTV/fotos')
    if len(faces) != 0:
        print(len(faces), ids)
        recognizer.train(faces, np.array(ids))
        recognizer.write('trainer.yml') 
    print("{0} faces trained.".format(len(np.unique(ids))))

recognizer = cv2.face.LBPHFaceRecognizer.create()

try:
    recognizer.read('trainer.yml')
    cascadePath = "CCTV/Cascades/haarcascade_frontalface_default.xml"
    faceCascade = cv2.CascadeClassifier(cascadePath)
    font = cv2.FONT_HERSHEY_SIMPLEX
except:
    pass



logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler()  
    ]
)



app = Flask(__name__, static_url_path='/CCTV')
app.secret_key = os.urandom(24)
sistema = SistemaGestion()
sistema.registrar_teclado()

@app.route("/train_test")
def train_test():
    train(recognizer)
    return {'status': "Training completed"}

@app.route("/video_feed/<int:cam_index>")
def video_feed(cam_index):
    if cam_index >= len(sistema.camaras):
        return "Cámara no disponible", 404
    return Response(sistema.video_feed(cam_index), mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/eliminar_miembro/<rut>", methods=["POST"])
def eliminar_miembro(rut):
    workbook = load_workbook(sistema.excel_file)
    sheet = workbook.active
    encontrado = False
    miembro = sistema.obtener_miembro(rut)
    try:
        folder_path_to_delete = os.path.join("CCTV/fotos", miembro[1].value.replace(".", ""))
        for file in os.listdir(folder_path_to_delete):
            os.remove(os.path.join(folder_path_to_delete, file))
    except:
        pass
    for row in sheet.iter_rows(min_row=2):
        if row[1].value == rut:  
            sheet.delete_rows(row[0].row)
            encontrado = True
            break

    if encontrado:
        workbook.save(sistema.excel_file)
        flash(f"El miembro con RUT {rut} ha sido eliminado exitosamente.")
    else:
        flash(f"No se encontró un miembro con el RUT {rut}.")

    return redirect(url_for("index"))

def apply_variations(image):
    """Aplica transformaciones y efectos a la imagen para generar variaciones."""
    variations = []

    # Rotación aleatoria
    angle = random.uniform(-15, 15)
    h, w = image.shape[:2]
    center = (w // 2, h // 2)
    rotation_matrix = cv2.getRotationMatrix2D(center, angle, 1.0)
    rotated = cv2.warpAffine(image, rotation_matrix, (w, h))
    variations.append(rotated)

    # Cambio de brillo y contraste aleatorio
    alpha = random.uniform(0.8, 1.2)  # Contraste
    beta = random.randint(-30, 30)    # Brillo
    bright_contrast = cv2.convertScaleAbs(image, alpha=alpha, beta=beta)
    variations.append(bright_contrast)

    # Desenfoque gaussiano
    kernel_size = random.choice([3, 5, 7])
    blurred = cv2.GaussianBlur(image, (kernel_size, kernel_size), 0)
    variations.append(blurred)

    # Agregar ruido aleatorio
    noise = np.random.normal(0, 25, image.shape).astype(np.uint8)
    noisy_image = cv2.add(image, noise)
    variations.append(noisy_image)

    # Cambios de color (solo si la imagen está en BGR)
    if len(image.shape) == 3:
        hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        hsv_image[..., 0] = (hsv_image[..., 0] + random.randint(-10, 10)) % 180
        color_shifted = cv2.cvtColor(hsv_image, cv2.COLOR_HSV2BGR)
        variations.append(color_shifted)

    return variations

def get_dataset(rut):
    face_detector = cv2.CascadeClassifier('CCTV/Cascades/haarcascade_frontalface_default.xml')
    count = 0
    while(True):
        # Abro la imagen con open CV para operar con ella
        img = cv2.imread(f'CCTV/fotos/{rut.replace(".", "")}/{rut}_foto.jpg')
        img = cv2.flip(img, 1) # 1 stright, 0 reverse
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_detector.detectMultiScale(gray, 1.3, 5)
        for (x,y,w,h) in faces:
            cv2.rectangle(img, (x,y), (x+w,y+h), (255,0,0), 2)    
            face = gray[y:y+h,x:x+w] 
            count += 1
            cv2.imwrite(f'CCTV/fotos/{rut.replace(".", "")}/{rut}_foto.{count}.jpg', face)
            
            # Aplicar variaciones y guardarlas
            variations = apply_variations(face)
            for i, variation in enumerate(variations, start=1):
                cv2.imwrite(f'CCTV/fotos/{rut.replace(".", "")}/{rut}_foto.{count}_var{i}.jpg', variation)
            #cv2.imshow('image', img)
        k = cv2.waitKey(100) & 0xff
        if k == 27:
            break
        elif count >= 100: # Take 100 face sample and stop video
            break
        


@app.route("/agregar_miembro", methods=["POST"])
def agregar_miembro():
    nombre = request.form.get("name") 
    rut = request.form.get("rut") 
    patente_input = request.form.get("patente")  
    rol = request.form.get("rol") 
    foto = request.files["foto"]
    
    sistema.agregar_miembro(nombre, rut, foto, patente_input, rol)
    get_dataset(rut)
    return redirect(url_for("index"))


@app.route("/registrar_salida/<rut>")
def registrar_salida(rut):
    sistema.registrar_salida(rut)
    return redirect(url_for("index"))

@app.route("/capture/<int:cam_index>")
def capture(cam_index):
    filename = sistema.capturar_imagen(cam_index)
    if filename:
        return jsonify({"message": "Imagen capturada y procesada con éxito", "file": filename}), 200
    return jsonify({"error": "No se pudo capturar la imagen"}), 500

@app.route("/recognition/<int:rut>")
def recognition(rut: int):
    info = sistema.obtener_miembro(rut)
    return {'photo_path': info}

def capture_by_frames(cam_index): 
    global cam
    cam = cv2.VideoCapture(cam_index)

    # SI CAM INDEX ES 0 o 1, SERÁ PARA RECONOCIMIENTO
    while True:
        try:
            ret, img = cam.read()
            if cam_index == 0 or cam_index == 1:
                img = cv2.flip(img, 1) # 1 Stright 0 Reverse
                img = cv2.resize(img, (400, 400))
                gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
                detector=cv2.CascadeClassifier(cascadePath)
                faces=detector.detectMultiScale(img,1.2,6)
                for(x,y,w,h) in faces:
                    cv2.rectangle(img, (x,y), (x+w,y+h), (0,255,0), 2)
                    id, confidence = recognizer.predict(gray[y:y+h,x:x+w])       

                    if (confidence < 100):
                        miembro = sistema.obtener_miembro(id, True)
                        id = miembro[0].value
                        confidence = "  {0}%".format(round(confidence))
                    else:
                        id = "Unknown"
                        confidence = "  {0}%".format(round(confidence))
                    p_count = photos_count()
                    if p_count != 0:
                        cv2.putText(img,f'{str(id)} {confidence}',(x+5,y-5),font,1,(255,255,255),2)
                    else:
                        cv2.putText(img,str('No Users Found'),(x+5,y-5),font,1,(255,255,255),2)

                    #cv2.putText(img,str(confidence),(x+5,y+h),font,1,(255,255,0),1)
            
            ret1, buffer = cv2.imencode('.jpg', img)
            frame = buffer.tobytes()
            
            
            yield (b'--frame\r\n'
                b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
        except Exception as e:
            cam.release()
            #break


@app.route('/video_capture_pro/<int:cam_index>')
def video_capture_pro(cam_index):
    print("CAM INDEX: ", cam_index)
    return Response(capture_by_frames(cam_index = cam_index), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/start',methods=['POST'])
def start():
    return render_template('index.html')

@app.route("/")
def index():
    registros = []
    workbook = load_workbook(sistema.excel_file)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):  
        registros.append(row)

    cantidad_camaras = len(sistema.camaras)
    registros_camaras = [cam is not None for cam in sistema.camaras]
    print(f"Cantidad de cámaras disponibles: {cantidad_camaras}")
    return render_template(
        "index.html",
        registros=registros,
        cantidad_camaras=cantidad_camaras,
        registros_camaras=registros_camaras
    )



def capture_frame(): 
    try:
        ret, img = cam.read()
        img = cv2.flip(img, 1) # 1 Stright 0 Reverse
        gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        detector=cv2.CascadeClassifier(cascadePath)
        faces=detector.detectMultiScale(img,1.2,6)
        for(x,y,w,h) in faces:
            cv2.rectangle(img, (x,y), (x+w,y+h), (0,255,0), 2)
            id, confidence = recognizer.predict(gray[y:y+h,x:x+w])
            if (confidence < 100):
                p_count = photos_count()
                if p_count != 0:
                    id = id
                    confidence = "  {0}%".format(round(100 - confidence))
                else:
                    id = "Unknown"
                    confidence = "  {0}%".format(round(100 - confidence))
            else:
                id = "Unknown"
                confidence = "  {0}%".format(round(100 - confidence))
            return {'id': id, 'confidence': confidence}
    except Exception as e:
        cam.release()
        print(e)
        return {'id': 'Unknown', 'confidence': '0%'}
    

def capture_frame_qr(): 
    try:
        qreader = QReader()
        ret, img = cam.read()
        img = cv2.flip(img, 1) # 1 Stright 0 Reverse
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        decoded_text = qreader.detect_and_decode(image=img)
        
        if decoded_text:
            return {'qr_content': str(decoded_text[0])}
        else:
            return {'qr_content': 'error'}
    except Exception as e:
        cam.release()
        print(e)
        return {'qr_content': 'error'}




@app.route("/capture_face")
def capture_face():
    predict = capture_frame()
    if predict['id'] != 'Unknown':
        miembro = sistema.obtener_miembro(predict['id'], True)
        return {'predict': predict, 'name': miembro[0].value, 'patente': miembro[3].value, 'rol': miembro[6].value, 'rut': miembro[1].value}
    return {'status': 'error'}



@app.route("/capture_qr")
def capture_qr():
    qr_content  = capture_frame_qr()
    if qr_content:
        return {'qr_content': qr_content}
    return {'status': 'error'}



@app.route("/start_engine/<int:motor_num>")
def start_engine(motor_num):
    if motor_num == '1':
        sistema.control_motor(1)
    elif motor_num == '2':
        sistema.control_motor(2)
    
    return {'status': 'ok'}

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
