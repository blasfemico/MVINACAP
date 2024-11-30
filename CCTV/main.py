import os
import time
import re
from datetime import datetime
from typing import List, Generator
import cv2
import easyocr
import requests
from flask import Flask, render_template, Response, request, redirect, url_for, jsonify, flash
from openpyxl import Workbook, load_workbook
import pytesseract
import numpy as np 
from pynput import keyboard
from typing import Tuple
import logging

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler()  
    ]
)

class SistemaGestion:
    def __init__(self):
        self.excel_file = "registro_miembros.xlsx"
        self.upload_folder = "/home/dante/Documentos/repositorio-git/MVINACAP/CCTV/fotos"
        self.esp32_motor_url = "http://192.168.100.49/motor"
        self.camaras = []
        self.latest_frames = []
        self.capturing = []
        self.reader = easyocr.Reader(["en"], gpu=True)
        self.setup()

    def setup(self):
        logging.info("Configurando el sistema...")
        if not os.path.exists(self.upload_folder):
            os.makedirs(self.upload_folder)
            logging.debug(f"Carpeta de subida creada: {self.upload_folder}")
        else:
            logging.debug(f"Carpeta de subida ya existe: {self.upload_folder}")

        if not os.path.exists(self.excel_file):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Nombre", "RUT", "Foto", "Patente", "Hora Entrada", "Hora Salida", "Vehiculo"])
            workbook.save(self.excel_file)
            logging.info(f"Archivo Excel creado: {self.excel_file}")
        else:
            logging.info(f"Archivo Excel ya existe: {self.excel_file}")
        
        pytesseract.pytesseract.tesseract_cmd = (
            "/usr/bin/tesseract" if os.name == "posix" else "C:/Program Files (x86)/Tesseract-OCR/tesseract.exe"
        )
        self.inicializar_camaras()

    def control_motor(self, motor_id):
        response = requests.get(f"{self.esp32_motor_url}/{motor_id}")
        response.raise_for_status()

    def capturar_imagen(self, cam_index):
        logging.info(f"Capturando imagen desde la cámara {cam_index}...")
        temp_folder = "/home/dante/Documentos/repositorio-git/MVINACAP/temp"
        if not os.path.exists(temp_folder):
            os.makedirs(temp_folder)
            logging.debug(f"Carpeta temporal creada: {temp_folder}")

        if cam_index < len(self.camaras):
            success, frame = self.camaras[cam_index].read()
            if success:
                image_path = os.path.join(temp_folder, f"captura_camara_{cam_index}_{int(time.time())}.jpg")
                cv2.imwrite(image_path, frame)
                logging.info(f"Imagen capturada: {image_path}")
                self.aplicar_filtros(image_path)
                matriculas_detectadas = self.detectar_matricula(image_path)
                if not matriculas_detectadas:
                    logging.warning("No se detectaron matrículas en la imagen capturada.")
                    return image_path
                saved_paths = self.guardar_matriculas(image_path, matriculas_detectadas)
                if not saved_paths:
                    logging.warning("No se pudieron guardar las matrículas detectadas.")
                    return image_path

                for saved_path in saved_paths:
                    matricula = self.leer_matricula(saved_path)
                    if not matricula:
                        logging.warning(f"No se pudo leer ninguna matrícula en la imagen recortada: {saved_path}.")
                        continue
                    es_valida, matricula_registrada = self.verificar_matricula(matricula)
                    if es_valida:
                        logging.info(f"Matrícula válida detectada: {matricula_registrada}.")
                        self.registrar_hora_entrada(matricula_registrada)
                    else:
                        logging.warning(f"Matrícula {matricula} no válida o no registrada.")
                return image_path
            else:
                logging.error(f"Error al capturar imagen desde la cámara {cam_index}")
        else:
            logging.warning(f"Cámara {cam_index} no está disponible.")
        return None

        
    def aplicar_filtros(self, image_path):
        logging.info(f"Aplicando filtros a la imagen: {image_path}")
        img = cv2.imread(image_path)
        if img is None:
            logging.error(f"Error: No se pudo abrir la imagen en la ruta especificada: {image_path}")
            return

        output_dir = os.path.join("/home/dante/Documentos/repositorio-git/MVINACAP/temp/output")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logging.debug(f"Directorio de salida creado: {output_dir}")

        try:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)
            edges = cv2.Canny(blurred, 50, 200)

            cv2.imwrite(os.path.join(output_dir, "image_gray.jpg"), gray)
            cv2.imwrite(os.path.join(output_dir, "image_blurred.jpg"), blurred)
            cv2.imwrite(os.path.join(output_dir, "image_edges.jpg"), edges)
            logging.info("Filtros aplicados y guardados.")
        except Exception as e:
            logging.error(f"Error aplicando filtros: {e}")


    def capturar_frame(self, cam_index):
        if cam_index < len(self.camaras) and self.camaras[cam_index] is not None:
            success, frame = self.camaras[cam_index].read()
            if success:
                self.latest_frames[cam_index] = frame

    def video_feed(self, cam_index) -> Generator[bytes, None, None]:
        while True:
            if cam_index < len(self.camaras) and self.camaras[cam_index] is not None:
                self.capturar_frame(cam_index)
                if self.latest_frames[cam_index] is not None:
                    _, buffer = cv2.imencode(".jpg", self.latest_frames[cam_index])
                    yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + buffer.tobytes() + b"\r\n"
            else:
                empty_frame = cv2.imread("fotos/cameraoff.jpg")
                if empty_frame is None:
                    empty_frame = self.generar_frame_azul()
                _, buffer = cv2.imencode(".jpg", empty_frame)
                yield b"--frame\r\nContent-Type: image/jpeg\r\n\r\n" + buffer.tobytes() + b"\r\n"
                time.sleep(0.1)

    def generar_frame_azul(self):
        """Genera un frame azul como placeholder."""
        azul_frame = cv2.imread("fotos/blue_screen.jpg") 
        if azul_frame is None:
            azul_frame = 255 * cv2.ones((480, 640, 3), dtype=np.uint8)
        return azul_frame


    def inicializar_camaras(self):
        max_camaras = 4
        for index in range(max_camaras):
            cam = cv2.VideoCapture(index)
            if cam.isOpened():
                self.camaras.append(cam)
                self.latest_frames.append(None)
                self.capturing.append(False)
                print(f"Cámara {index} inicializada correctamente.")
            else:
                self.camaras.append(None)  
                self.latest_frames.append(None)
                self.capturing.append(False)
                print(f"Advertencia: No se pudo inicializar la cámara {index}. Ignorando...")
        if not any(self.camaras):
            print("No se encontraron cámaras disponibles.")

    def detectar_matricula(self, image_path) -> List[str]:
        logging.info(f"Detectando matrícula en la imagen: {image_path}")
        img = cv2.imread(image_path)
        if img is None:
            logging.error(f"Error: No se pudo abrir la imagen en la ruta especificada: {image_path}")
            return []

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blurred, 50, 150)
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        matriculas = []
        for contour in contours:
            area = cv2.contourArea(contour)
            if area > 5000:  
                peri = cv2.arcLength(contour, True)
                approx = cv2.approxPolyDP(contour, 0.02 * peri, True)
                if len(approx) == 4:  
                    matriculas.append(approx)
                    logging.debug(f"Posible matrícula detectada con área: {area}")

        if not matriculas:
            logging.warning("No se detectaron matrículas en la imagen.")
        return matriculas

    def guardar_matriculas(self, image_path, matriculas) -> List[str]:
        img = cv2.imread(image_path)
        rutas_matriculas = []
        temp_folder = "/home/dante/Documentos/repositorio-git/MVINACAP/temp"
        if not os.path.exists(temp_folder):
            os.makedirs(temp_folder)
            logging.debug(f"Carpeta temporal creada: {temp_folder}")
        for i, matricula in enumerate(matriculas):
            x, y, w, h = cv2.boundingRect(matricula)
            matricula_recortada = img[y:y + h, x:x + w]
            ruta = os.path.join(temp_folder, f"matricula_{i+1}.jpg")
            cv2.imwrite(ruta, matricula_recortada)
            rutas_matriculas.append(ruta)
        return rutas_matriculas


    def leer_matricula(self, image_path) -> str:
        logging.info(f"Intentando leer matrícula en la imagen: {image_path}")
        if not os.path.exists(image_path):
            logging.error(f"Error: No se encontró la imagen en la ruta especificada: {image_path}")
            return ""

        image = cv2.imread(image_path)
        if image is None:
            logging.error(f"Error: No se pudo abrir la imagen en la ruta especificada: {image_path}")
            return ""

        image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        result = self.reader.readtext(image_rgb)

        if not result:
            logging.warning("No se detectó ningún texto en la imagen.")
            return ""

        for bbox, text, prob in result:
            logging.debug(f"Texto detectado: '{text}' con probabilidad: {prob:.2f}")
            if prob > 0.55:
                logging.info(f"Matrícula detectada: {text}")
                return text

        logging.warning("No se detectó ninguna matrícula confiable en la imagen.")
        return ""



    def verificar_matricula(self, matricula: str) -> Tuple[bool, str]:
        logging.info(f"Verificando matrícula: {matricula}")
        workbook = load_workbook(self.excel_file)
        sheet = workbook.active
        matriculas_registradas = [row[3].value for row in sheet.iter_rows(min_row=2) if row[3].value]

        for registrada in matriculas_registradas:
            similitud = self.similitud_matriculas(matricula, registrada)
            if similitud > 0.8:
                logging.info(f"Matrícula '{matricula}' coincide con '{registrada}' (Similitud: {similitud:.2f})")
                return True, registrada

        logging.warning(f"Matrícula '{matricula}' no coincide con ninguna registrada.")
        return False, None


    def similitud_matriculas(self, matricula1: str, matricula2: str) -> float:
        from difflib import SequenceMatcher
        return SequenceMatcher(None, matricula1, matricula2).ratio()

    def registrar_hora_entrada(self, matricula):
        hora_entrada = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        workbook = load_workbook(self.excel_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            if row[3].value == matricula and row[4].value is None:
                row[4].value = hora_entrada
                workbook.save(self.excel_file)
                logging.info(f"Hora de entrada registrada para la matrícula {matricula}: {hora_entrada}.")
                #motor 1
                self.control_motor(1)
                logging.info(f"Motor 1 activado para la matrícula {matricula}.")
                return
        logging.warning(f"No se encontró una fila para registrar la hora de entrada para la matrícula {matricula}.")


    def procesar_matriculas(self, image_paths: List[str], motor_id: int):
        for image_path in image_paths:
            logging.info(f"Procesando imagen: {image_path}")
            matriculas_detectadas = self.detectar_matricula(image_path)
            if not matriculas_detectadas:
                logging.warning(f"No se detectaron matrículas en la imagen: {image_path}")
                continue

            saved_paths = self.guardar_matriculas(image_path, matriculas_detectadas)
            if not saved_paths:
                logging.warning(f"No se generaron imágenes recortadas de matrículas para: {image_path}")
                continue

            for saved_path in saved_paths:
                matricula = self.leer_matricula(saved_path)
                if matricula:
                    es_valida, matricula_registrada = self.verificar_matricula(matricula)
                    if es_valida:
                        if motor_id == 1:
                            self.registrar_hora_entrada(matricula_registrada)
                        elif motor_id == 2:
                            self.registrar_hora_salida(matricula_registrada)
                        self.control_motor(motor_id)
                        logging.info(f"Motor {motor_id} activado para la matrícula {matricula}.")
                    else:
                        logging.warning(f"No se activó el motor. Matrícula {matricula} no es válida.")
                else:
                    logging.warning(f"No se pudo leer ninguna matrícula en la imagen recortada: {saved_path}")


    def limpiar_matricula(matricula: str) -> str:
        matricula_limpia = re.sub(r"[^A-Za-z0-9]", "", matricula)
        return matricula_limpia

    def agregar_miembro(self, nombre, rut, foto, patente_input, vehiculo):
        foto_path = os.path.join(self.upload_folder, f"{rut}_foto.jpg")
        if not os.path.exists(self.upload_folder):
            os.makedirs(self.upload_folder)
            logging.debug(f"Carpeta de fotos creada: {self.upload_folder}")
        foto.save(foto_path)
        logging.info(f"Foto guardada en: {foto_path}")
        workbook = load_workbook(self.excel_file)
        sheet = workbook.active
        sheet.append([nombre, rut, foto_path, patente_input, None, vehiculo])
        workbook.save(self.excel_file)
        logging.info(f"Datos del miembro agregados al Excel: Nombre={nombre}, RUT={rut}, Patente={patente_input}, Vehículo={vehiculo}")
        
        flash("Miembro agregado exitosamente.", "success")



    def registrar_hora_salida(self, matricula):
        hora_salida = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        workbook = load_workbook(self.excel_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            if row[3].value == matricula and row[5].value is None:
                row[5].value = hora_salida
                workbook.save(self.excel_file)
                logging.info(f"Hora de salida registrada para la matrícula {matricula}: {hora_salida}.")
                #motor 2
                self.control_motor(2)
                logging.info(f"Motor 2 activado para la matrícula {matricula}.")
                return
        logging.warning(f"No se encontró una fila para registrar la hora de salida para la matrícula {matricula}.")


    def registrar_teclado(self):
        def on_press(key):
            try:
                if key.char == 'q': 
                    print("Capturando imágenes desde cámaras 0 y 1 para entrada...")
                    capturas = self.capturar_imagenes([0, 1])
                    self.procesar_matriculas(capturas, motor_id=1)
                elif key.char == 'w':  
                    print("Capturando imágenes desde cámaras 2 y 3 para salida...")
                    capturas = self.capturar_imagenes([2, 3])
                    self.procesar_matriculas(capturas, motor_id=2)
            except AttributeError:
                pass

        listener = keyboard.Listener(on_press=on_press)
        listener.start()

    
    def activar_motor_entrada(self):
        self.control_motor(1)  

    def activar_motor_salida(self):
        self.control_motor(2)  
        

    def obtener_registros(self):
        registros = []
        if os.path.exists(self.excel_file):
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                registros.append(row)
        else:
            print("Advertencia: El archivo de registros no existe.")
        return registros

app = Flask(__name__, static_url_path='/CCTV', static_folder='/home/dante/Documentos/repositorio-git/MVINACAP/CCTV')
app.secret_key = os.urandom(24)
sistema = SistemaGestion()
sistema.registrar_teclado()

@app.route("/video_feed/<int:cam_index>")
def video_feed(cam_index):
    if cam_index >= len(sistema.camaras):
        return "Cámara no disponible", 404
    return Response(sistema.video_feed(cam_index), mimetype="multipart/x-mixed-replace; boundary=frame")

@app.route("/eliminar_miembro/<rut>", methods=["POST"])
def eliminar_miembro(rut):
    workbook = load_workbook(sistema.excel_file)
    sheet = workbook.active
    encontrado = False
    for row in sheet.iter_rows(min_row=2):
        if row[1].value == rut:  
            sheet.delete_rows(row[0].row)
            encontrado = True
            break

    if encontrado:
        workbook.save(sistema.excel_file)
        flash(f"El miembro con RUT {rut} ha sido eliminado exitosamente.")
    else:
        flash(f"No se encontró un miembro con el RUT {rut}.")
    return redirect(url_for("index"))



@app.route("/agregar_miembro", methods=["POST"])
def agregar_miembro():
    nombre = request.form.get("name") 
    rut = request.form.get("rut") 
    patente_input = request.form.get("patente")  
    vehiculo = request.form.get("vehiculo") 
    foto = request.files["foto"]  
    sistema.agregar_miembro(nombre, rut, foto, patente_input, vehiculo)
    return redirect(url_for("index"))


@app.route("/registrar_salida/<rut>")
def registrar_salida(rut):
    sistema.registrar_salida(rut)
    return redirect(url_for("index"))

@app.route("/capture/<int:cam_index>")
def capture(cam_index):
    filename = sistema.capturar_imagen(cam_index)
    if filename:
        return jsonify({"message": "Imagen capturada y procesada con éxito", "file": filename}), 200
    return jsonify({"error": "No se pudo capturar la imagen"}), 500

@app.route("/")
def index():
    registros = sistema.obtener_registros()
    cantidad_camaras = len(sistema.camaras)
    registros_camaras = [cam is not None for cam in sistema.camaras]
    print(f"Cantidad de cámaras disponibles: {cantidad_camaras}")
    return render_template(
        "index.html",
        registros=registros,
        cantidad_camaras=cantidad_camaras,
        registros_camaras=registros_camaras
    )

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
