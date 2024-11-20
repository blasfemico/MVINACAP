import time
from werkzeug.utils import secure_filename
import re
from datetime import datetime
from typing import List, Generator
import os
from flask import (
    Flask,
    render_template,
    Response,
    request,
    redirect,
    url_for,
    jsonify,
    send_from_directory,
    flash,
)
import cv2
import easyocr
import openpyxl
import requests
from openpyxl import load_workbook
import pytesseract
import sys
print(sys.path)  # Muestra todas las rutas donde Python busca librería


EXCEL_FILE = "registro_miembros.xlsx"
UPLOAD_FOLDER = "C:/Users/Darkm/Desktop/CCTV/temp/"
ESP32_IP = "http://192.168.119.214/motor"
CAP = cv2.VideoCapture(1)
LATEST_FRAME = None
CAPTURING = False
LAST_DETECTED_PLATE= None
app = Flask(__name__)
app.secret_key = os.urandom(24)

if os.name == "posix":
    pytesseract.pytesseract.tesseract_cmd = "/usr/bin/tesseract"
elif os.name == "nt":
    pytesseract.pytesseract.tesseract_cmd = (
        "C:/Program Files (x86)/Tesseract-OCR/tesseract.exe"
    )

try:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    print(f"Directorio '{UPLOAD_FOLDER}' creado o ya existe.")
except OSError as e:
    print(f"Error al crear el directorio '{UPLOAD_FOLDER}': {e}")

if not os.path.exists(EXCEL_FILE):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(
        ["Nombre", "RUT", "Foto", "Patente", "Hora Entrada", "Hora Salida", "Vehiculo"]
    )
    workbook.save(EXCEL_FILE)


@app.route("/control_motor", methods=["POST"])
def control_motor():
    try:
        response = requests.get(ESP32_IP)
        response.raise_for_status()
        return jsonify({"message": "Motor controlado exitosamente"}), 200
    except requests.exceptions.ConnectionError:
        return (
            jsonify(
                {
                    "error": "No se pudo conectar al ESP32. Verifica la dirección IP y la conexión."
                }
            ),
            503,
        )
    except requests.exceptions.Timeout:
        return (
            jsonify(
                {"error": "La solicitud al ESP32 ha excedido el tiempo de espera."}
            ),
            504,
        )
    except requests.exceptions.RequestException as e:
        return (
            jsonify(
                {"error": f"Ocurrió un error al comunicarse con el ESP32: {str(e)}"}
            ),
            500,
        )


def detectar_matricula(imagen) -> List:
    img = cv2.imread(imagen)
    if img is None:
        print(f"Error: No se pudo abrir la imagen en la ruta especificada: {imagen}")
        return []

    # Convertir a escala de grises y aplicar filtros
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 13, 15, 15)  # Ajustar para reducir el ruido y preservar los bordes
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)

    # Detección de bordes mejorada
    edges = cv2.Canny(blurred, 50, 150)  # Ajuste de umbrales de Canny
    edges = cv2.dilate(edges, None, iterations=1)  # Dilatar los bordes para mejorar los contornos
    edges = cv2.erode(edges, None, iterations=1)  # Erosionar para refinar los bordes

    # Encontrar contornos
    contours, _ = cv2.findContours(edges.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    print(f"Contornos encontrados: {len(contours)}")

    matriculas = []
    for contour in contours:
        area = cv2.contourArea(contour)
        if 5000 < area < 30000:  # Ajuste de área para excluir contornos muy pequeños o muy grandes
            peri = cv2.arcLength(contour, True)
            approx = cv2.approxPolyDP(contour, 0.018 * peri, True)  # Mayor precisión en los contornos
            if len(approx) == 4:
                # Verificar proporción de aspecto del contorno (ancho:alto)
                (x, y, w, h) = cv2.boundingRect(approx)
                aspect_ratio = w / float(h)
                if 2 < aspect_ratio < 5:  # Ajuste de la relación de aspecto típico de una matrícula
                    matriculas.append(approx)
                    cv2.drawContours(img, [approx], -1, (0, 255, 0), 3)

    # Guardar la imagen con contornos de matrículas detectadas
    cv2.imwrite("output/image_contours.jpg", img)
    return matriculas


def guardar_matriculas(imagen, matriculas) -> List[str]:
    img = cv2.imread(imagen)
    rutas_matriculas = []
    for i in range(len(matriculas)):
        ruta = f"{UPLOAD_FOLDER}matricula_completa{i+1}.jpg"
        cv2.imwrite(ruta, img)
        rutas_matriculas.append(ruta)

    return rutas_matriculas


def leer_matricula(image_path) -> str:
    reader = easyocr.Reader(["en"])
    if not os.path.exists(image_path):
        print(f"Error: No se encontró la imagen en la ruta especificada: {image_path}")
        return ""
    image = cv2.imread(image_path)
    if image is None:
        print(
            f"Error: No se pudo abrir la imagen en la ruta especificada: {image_path}"
        )
        return ""
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    result = reader.readtext(image_rgb)
    matricula = ""
    for bbox, text, prob in result:
        print(f"Detectado: {text} con una probabilidad de {prob:.2f}")
        if matricula == "" or prob > 0.69:
            matricula = text
    print(f"Matrícula: {matricula}")
    return matricula


def limpiar_matricula(matricula) -> str:
    matricula_limpia = re.sub(r"[^A-Za-z0-9]", "", matricula)
    return matricula_limpia


def gen_frames() -> Generator[bytes, None, None]:
    global CAP, LATEST_FRAME, CAPTURING
    while True:
        while CAPTURING:
            time.sleep(0.1)

        if not CAP.isOpened():
            print("Reconectando a la cámara...")
            CAP.release()
            CAP.open(URL)

        success, frame = CAP.read()
        if not success:
            print("Error: No se pudo capturar la imagen.")
            time.sleep(2)
            continue

        LATEST_FRAME = frame

        _, buffer = cv2.imencode(".jpg", LATEST_FRAME)
        frame_bytes = buffer.tobytes()
        yield (
            b"--frame\r\n" b"Content-Type: image/jpeg\r\n\r\n" + frame_bytes + b"\r\n"
        )


@app.route("/video_feed")
def video_feed() -> Response:
    return Response(gen_frames(), mimetype="multipart/x-mixed-replace; boundary=frame")


@app.route("/fotos/<filename>")
def fotos(filename) -> Response:
    return send_from_directory("fotos", filename)


@app.route("/")
def index() -> str:
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    registros = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        registros.append(row)
    return render_template("index.html", registros=registros)


@app.route("/agregar_miembro", methods=["POST"])
def agregar_miembro() -> str:
    nombre = request.form.get("name")
    rut = request.form.get("rut")
    patente_input = request.form.get("patente")
    vehiculo = request.form.get("vehiculo")
    foto = request.files["foto"]

    if nombre and rut and foto and patente_input:
        foto_filename = os.path.join("fotos/", f"{rut}_foto.jpg")
        foto.save(foto_filename)

        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

        # Agregar los datos sin registrar la hora de entrada
        sheet.append(
            [
                nombre,
                rut,
                foto_filename,
                patente_input,
                None,       # Dejar la columna de hora de entrada vacía
                None,       # Columna de hora de salida vacía
                vehiculo,
            ]
        )

        workbook.save(EXCEL_FILE)
        flash("Datos agregados correctamente", "success")
        return redirect(url_for("index"))

    flash("Faltan datos o la foto no se subió. Por favor complete todos los campos.", "error")
    return redirect(url_for("index"))



@app.route("/editar_miembro/<rut>", methods=["GET", "POST"])
def editar_miembro(rut) -> str:
    print(f"Editando miembro con RUT: {rut}")
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    if request.method == "POST":
        nombre = request.form.get("name")
        patente = request.form.get("patente")
        vehiculo = request.form.get("vehiculo")
        print(
            f"Datos recibidos: nombre={nombre}, patente={patente}, vehiculo={vehiculo}"
        )
        for row in sheet.iter_rows(min_row=2):
            if row[1].value == rut:
                print(f"Modificando el miembro: {row[0].value} con RUT {rut}")
                row[0].value = nombre
                row[3].value = patente
                row[6].value = vehiculo
                workbook.save(EXCEL_FILE)
                return redirect(url_for("index"))
    else:
        miembro = None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == rut:
                miembro = row
                break
        if miembro is None:
            print(f"No se encontró miembro con RUT: {rut}")
        return render_template("editar_miembro.html", miembro=miembro)


@app.route("/eliminar_miembro/<rut>", methods=["GET"])
def eliminar_miembro(rut) -> str:
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    ruta = f"fotos/{rut}_foto.jpg"
    for row in sheet.iter_rows(min_row=2):
        if row[1].value == rut:
            sheet.delete_rows(row[0].row)
            workbook.save(EXCEL_FILE)
            break
    if os.path.exists(ruta):
        os.remove(ruta)
        print(f"Archivo {ruta} eliminado.")
    else:
        print(f"El archivo {ruta} no existe.")
    return redirect(url_for("index"))


@app.route("/capture")
def capture():
    global CAPTURING, LATEST_FRAME, LAST_DETECTED_PLATE
    CAPTURING = True
    time.sleep(1)

    if LATEST_FRAME is not None:
        captured_image_path = os.path.join(UPLOAD_FOLDER, "captura.jpg")
        cv2.imwrite(captured_image_path, LATEST_FRAME)
        print("Imagen capturada.")

        matriculas_detectadas = detectar_matricula(captured_image_path)
        if matriculas_detectadas:
            rutas_matriculas = guardar_matriculas(captured_image_path, matriculas_detectadas)
            if rutas_matriculas:
                # Leer y limpiar la matrícula detectada
                detected_plate = leer_matricula(rutas_matriculas[0])
                detected_plate = limpiar_matricula(detected_plate)
                print(f"Matrícula detectada y procesada: {detected_plate}")

                # Cargar el archivo Excel y buscar coincidencias
                workbook = load_workbook(EXCEL_FILE)
                sheet = workbook.active
                datos_encontrados = False
                hora_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Buscar si la matrícula detectada coincide con alguna en el Excel
                for row in sheet.iter_rows(min_row=2, values_only=False):
                    excel_plate = row[3].value  # Columna de la matrícula en Excel
                    if excel_plate == detected_plate:  # Comparar con la matrícula detectada
                        datos_encontrados = True
                        LAST_DETECTED_PLATE = excel_plate  # Guardar la matrícula detectada en Excel

                        # Actualizar siempre la hora de entrada al detectar una nueva entrada
                        row[4].value = hora_actual  # Actualiza la hora de entrada
                        row[5].value = None         # Limpia la hora de salida (opcional, si quieres reiniciarla)
                        break

                # Mostrar error si no se encontró la matrícula en el Excel
                if not datos_encontrados:
                    print("Error: Matrícula no encontrada en el archivo Excel.")
                    flash("Error: Matrícula no encontrada en el archivo Excel.", "error")
                    CAPTURING = False
                    return redirect(url_for("index"))

                # Guardar cambios en el Excel y activar el motor
                workbook.save(EXCEL_FILE)
                print("Comando 'GIRO' enviado al ESP32.")
                control_motor()
                flash("Datos actualizados correctamente.", "success")
                CAPTURING = False
                return redirect(url_for("index"))
            else:
                print("Error: No se pudieron guardar las matrículas recortadas.")
                LAST_DETECTED_PLATE = None
                flash("Error: No se pudieron guardar las matrículas recortadas.", "error")
                return redirect(url_for("index"))
        else:
            print("No se detectó ninguna matrícula en la imagen capturada.")
            LAST_DETECTED_PLATE = None
            flash("No se detectó ninguna matrícula en la imagen capturada.", "error")
            return redirect(url_for("index"))
    else:
        print("Error: No se pudo capturar la imagen.")
        LAST_DETECTED_PLATE = None
    CAPTURING = False
    flash("Imagen capturada, pero no se detectó ninguna matrícula.", "error")
    return redirect(url_for("index"))

@app.route("/close_camera")
def close_camera() -> str:
    CAP.release()
    print("Cámara cerrada.")
    return "Cámara cerrada."


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)